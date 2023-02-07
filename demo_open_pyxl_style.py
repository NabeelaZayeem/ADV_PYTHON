from openpyxl.styles import Border,Side,Font,PatternFill,Alignment
from openpyxl import Workbook
wb=Workbook()
sheet=wb.active
sheet.merge_cells('A1:D1')
cell=sheet['B2']
cell.value='HCL'
thin=Side(border_style="thin",color="000000")
double=Side(border_style='double',color='ff0000')
cell.border=Border(top=thin,bottom=double)
cell.fill=PatternFill("solid",fgColor='DDDDDD')
cell.font=Font(b=True,color='FF0000')
cell.alignment=Alignment(horizontal='center',vertical='center')
wb.save('C://Users/user794/Documents/openpyxlstyle.xlsx')