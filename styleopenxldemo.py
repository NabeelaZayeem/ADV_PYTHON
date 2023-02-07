from openpyxl import Workbook
wb=Workbook()
sheet=wb.active
sheet.title='HCL'
# sheet['A1'].value=10
# sheet['B2'].value='Test'
# # sheet.cell(row=3,column=4).value='Hello'
# j=0
# for i in range(10,60,10):
#     j+=1
#     sheet.cell(row=j, column=i).value = i
#     sheet.cell(row=1,column=j).value=i #for printing 10 to 50 in a single row
# for row in sheet.iter_rows(min_row=1,max_row=5,min_col=2,max_col=2):
#     for r in row:
#             r.value=1


#######################################################

j=100
for col in sheet.iter_cols(min_row=1,max_row=1,min_col=1,max_col=5):
    j+=100
    for i in col:
        i.value=j
wb.save('C://Users/user794/Documents/demoopenxlwrite.xlsx')